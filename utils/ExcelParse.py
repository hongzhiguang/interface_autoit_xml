from openpyxl import load_workbook
import traceback
import os,time,sys
from config.ProjVar import *

class ParseExcel(object):

    def __init__(self,file_path):
        self.file_path = file_path
        self.wb = load_workbook(self.file_path)
        self.ws = self.setSheetByName(self.wb.sheetnames[0])

    def setSheetByName(self,name):
        """设置ws"""
        if name in self.wb.sheetnames:
            self.ws = self.wb[name]
            return self.ws
        self.ws = None
        return self.ws

    def getCols(self,rowNo):
        """指定行获取所有的列"""
        if not isinstance(rowNo,int):
            return None
        try:
            return list(self.ws.rows)[rowNo-1]
        except:
            traceback.print_exc()

    def getRows(self,colNo):
        """指定列获取所有的行"""
        if not isinstance(colNo,int):
            return None
        try:
            return list(self.ws.columns)[colNo-1]
        except:
            traceback.print_exc()

    def get_cell_obj(self, rowNo, colNo):
        """获取单元格对象"""
        if (not isinstance(rowNo, int)) or (not isinstance(colNo, int)):
            return None
        try:
            return self.ws.cell(row=rowNo, column=colNo)
        except:
            traceback.print_exc()

    def get_cell_value(self, rowNo, colNo):
        """参数行号和列表从1开始表示第一行"""
        if (not isinstance(rowNo, int)) or (not isinstance(colNo, int)):
            return None
        try:
            return self.ws.cell(row=rowNo, column=colNo).value
        except:
            traceback.print_exc()

    def get_min_row(self):
        # 获取表格中有数据的最小行数
        try:
            return self.ws.min_row
        except:
            traceback.print_exc()
            return None

    def get_max_row(self):
        # 获取表格中有数据的最大行数
        try:
            return self.ws.max_row
        except:
            traceback.print_exc()
            return None

    def get_min_column(self):
        # 获取表格中有数据的最小列数
        try:
            return self.ws.min_column
        except:
            traceback.print_exc()
            return None

    def get_max_column(self):
        # 获取表格中有数据的最大列数
        try:
            return self.ws.max_column
        except:
            traceback.print_exc()
            return None

    def write_cell(self, rowNo, colNo, content):
        """参数行号和列表从1开始表示第一行"""
        if (not isinstance(rowNo, int)) or (not isinstance(colNo, int)):
            return None
        try:
            self.ws.cell(row=rowNo, column=colNo).value = content
        except Exception:
            traceback.print_exc()

    def set_font(new_name, new_size, new_horizontal, new_vertical, new_bold=True):
        """设置单元格格式例子：ws.cell(1,1).font=set_font(...)"""
        from openpyxl.styles import Font, Alignment
        font = Font(name=new_name, size=new_size, bold=new_bold)
        align = Alignment(horizontal=new_horizontal, vertical=new_vertical)

    def save(self):
        # 表格中写入数据，保存生效
        try:
            self.wb.save(self.file_path)
        except PermissionError:
            print("PermissionError, Please close the file before saving.")
            sys.exit()

if __name__ == "__main__":
    excel = ParseExcel(testCasePath)